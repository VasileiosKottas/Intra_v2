# app/services/email_service.py
"""
SMTP Email Service for automated team reports
"""

import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta
from typing import Optional, Dict, List
import logging
import os
from app.services.report_service import ReportService

logger = logging.getLogger(__name__)

class SMTPEmailService:
    """Service for sending emails via SMTP"""
    
    def __init__(self, smtp_server: str, smtp_port: int, username: str, password: str, use_tls: bool = True):
        self.smtp_server = smtp_server
        self.smtp_port = smtp_port
        self.username = username
        self.password = password
        self.use_tls = use_tls
        self.report_service = ReportService()
        
    @classmethod
    def from_env(cls):
        """Create SMTP service from environment variables"""
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', '587'))
        username = os.getenv('SMTP_USERNAME', '')
        password = os.getenv('SMTP_PASSWORD', '')
        use_tls = os.getenv('SMTP_USE_TLS', 'true').lower() == 'true'
        
        if not username or not password:
            raise ValueError("SMTP_USERNAME and SMTP_PASSWORD must be set in environment variables")
            
        return cls(smtp_server, smtp_port, username, password, use_tls)
        
    def _create_email_body(self, team_name: str, report_month: str, team_metrics: Dict) -> str:
        """Create HTML email body for team report"""
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <style>
                body {{ 
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
                    line-height: 1.6; 
                    color: #333; 
                    margin: 0; 
                    padding: 0; 
                    background-color: #f4f4f4;
                }}
                .container {{ 
                    max-width: 800px; 
                    margin: 0 auto; 
                    background: white; 
                    border-radius: 10px; 
                    overflow: hidden; 
                    box-shadow: 0 0 20px rgba(0,0,0,0.1);
                }}
                .header {{ 
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    color: white; 
                    padding: 40px 30px; 
                    text-align: center; 
                }}
                .header h1 {{ 
                    margin: 0 0 10px 0; 
                    font-size: 32px; 
                    font-weight: 300; 
                }}
                .header p {{ 
                    margin: 0; 
                    font-size: 18px; 
                    opacity: 0.9; 
                }}
                .content {{ 
                    padding: 40px 30px; 
                }}
                .metrics-grid {{ 
                    display: grid; 
                    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); 
                    gap: 25px; 
                    margin: 30px 0; 
                }}
                .metric-card {{ 
                    background: #f8f9fa; 
                    padding: 25px 20px; 
                    border-radius: 12px; 
                    text-align: center; 
                    border: 2px solid #e9ecef; 
                    transition: all 0.3s ease;
                }}
                .metric-value {{ 
                    font-size: 28px; 
                    font-weight: bold; 
                    color: #667eea; 
                    margin-bottom: 8px; 
                    display: block;
                }}
                .metric-label {{ 
                    color: #6c757d; 
                    font-size: 14px; 
                    font-weight: 500;
                    text-transform: uppercase;
                    letter-spacing: 0.5px;
                }}
                .highlight {{ 
                    background: linear-gradient(135deg, #e3f2fd 0%, #bbdefb 100%); 
                    padding: 25px; 
                    border-left: 5px solid #2196f3; 
                    margin: 30px 0; 
                    border-radius: 0 12px 12px 0; 
                }}
                .highlight-icon {{ 
                    font-size: 24px; 
                    margin-right: 10px; 
                }}
                .footer {{ 
                    text-align: center; 
                    padding: 30px; 
                    background: #f8f9fa; 
                    color: #6c757d; 
                    font-size: 14px; 
                    border-top: 1px solid #e9ecef;
                }}
                .attachment-note {{ 
                    background: linear-gradient(135deg, #fff3cd 0%, #ffeaa7 100%); 
                    border: 2px solid #ffc107; 
                    padding: 25px; 
                    border-radius: 12px; 
                    margin: 30px 0; 
                }}
                .attachment-note strong {{ 
                    color: #856404; 
                    font-size: 18px;
                }}
                .attachment-note p {{ 
                    margin: 10px 0 0 0; 
                    color: #856404;
                }}
                .team-summary {{
                    background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
                    padding: 25px;
                    border-radius: 12px;
                    margin: 25px 0;
                }}
                .summary-title {{
                    font-size: 20px;
                    font-weight: 600;
                    color: #495057;
                    margin-bottom: 15px;
                    text-align: center;
                }}
                @media (max-width: 600px) {{
                    .header {{ padding: 30px 20px; }}
                    .header h1 {{ font-size: 28px; }}
                    .content {{ padding: 30px 20px; }}
                    .metrics-grid {{ grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 15px; }}
                    .metric-card {{ padding: 20px 15px; }}
                    .metric-value {{ font-size: 24px; }}
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>📊 {team_name} Weekly Performance Report</h1>
                    <p>Week of {report_month} | Automatically Generated</p>
                </div>
                
                <div class="content">
                    <div class="highlight">
                        <span class="highlight-icon">🎯</span>
                        <strong>Team Performance Summary</strong><br>
                        Your detailed Excel report is attached to this email with complete metrics, individual performance data, and goal tracking.
                    </div>
                    
                    <div class="team-summary">
                        <div class="summary-title">Key Performance Metrics</div>
                        <div class="metrics-grid">
                            <div class="metric-card">
                                <div class="metric-value">£{team_metrics.get('total_submitted', 0):,.0f}</div>
                                <div class="metric-label">Total Submitted</div>
                            </div>
                            <div class="metric-card">
                                <div class="metric-value">{team_metrics.get('total_applications', 0)}</div>
                                <div class="metric-label">Applications</div>
                            </div>
                            <div class="metric-card">
                                <div class="metric-value">{team_metrics.get('team_size', 0)}</div>
                                <div class="metric-label">Team Members</div>
                            </div>
                            <div class="metric-card">
                                <div class="metric-value">{team_metrics.get('conversion_rate', 0):.1f}%</div>
                                <div class="metric-label">Conversion Rate</div>
                            </div>
                        </div>
                    </div>
                    
                    <div class="attachment-note">
                        <strong>📎 Complete Excel Report Attached</strong>
                        <p>The attached Excel file contains detailed breakdowns including:</p>
                        <ul style="margin: 15px 0; padding-left: 25px; color: #856404;">
                            <li>Individual advisor performance metrics</li>
                            <li>Activity tracking (calls, appointments, applications)</li>
                            <li>Conversion rates and goal comparisons</li>
                            <li>Fee breakdowns and submission totals</li>
                        </ul>
                    </div>
                    
                    <p style="text-align: center; margin-top: 30px; color: #6c757d;">
                        Questions about this report? Contact your team administrator or visit the sales dashboard.
                    </p>
                </div>
                
                <div class="footer">
                    <p><strong>Sales Dashboard System</strong> | Automated Weekly Report</p>
                    <p>Report generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>
                    <p style="margin-top: 15px; font-size: 12px; color: #adb5bd;">
                        This email was sent automatically. Please do not reply to this message.
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
    
    def send_team_report_email(self, 
                              sender_email: str, 
                              recipient_emails: List[str], 
                              team_id: int, 
                              team_name: str,
                              subject_override: Optional[str] = None) -> bool:
        """Send team performance report email with Excel attachment"""
        
        try:
            # Determine which month to report on and create display text
            report_month, month_display = self._determine_report_month()
            
            # Get team data using the same method as YTD dashboard
            team_data = self._get_team_ytd_data(team_id)
            if not team_data or team_data.get('error'):
                logger.error(f"Failed to generate team data: {team_data.get('error') if team_data else 'No data returned'}")
                return False
            
            # Generate Excel using the YTD data
            excel_bytes = self._generate_excel_from_ytd_data(team_data, team_name, report_month)
            
            # Calculate team metrics for email summary using YTD data
            team_metrics = self._calculate_team_summary_from_ytd(team_data)
            
            # Create email content with proper month display
            subject = subject_override or f"📊 {team_name} Performance Report - {month_display}"
            html_body = self._create_email_body(team_name, month_display, team_metrics)
            
            # Create plain text version for better compatibility
            text_body = f"""
{team_name} Monthly Performance Report - {month_display}

Monthly Key Metrics:
- Total Monthly Submitted: £{team_metrics.get('total_submitted', 0):,.0f}
- Total Applications: {team_metrics.get('total_applications', 0)}
- Team Members: {team_metrics.get('team_size', 0)}
- Monthly Conversion Rate: {team_metrics.get('conversion_rate', 0):.1f}%

Your comprehensive monthly Excel report is attached to this email with complete monthly metrics and individual advisor performance data.

This is an automated monthly report generated by the Sales Dashboard System.
Report generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}
            """.strip()
            
            # Prepare Excel attachment with report month in filename
            filename = f"{team_name.replace(' ', '_')}_Performance_Report_{report_month.replace('-', '')}.xlsx"
            
            # Send email via SMTP
            return self._send_via_smtp(sender_email, recipient_emails, subject, text_body, html_body, excel_bytes, filename)
            
        except Exception as e:
            logger.error(f"Error sending team report email: {e}")
            return False
    
    def _get_team_ytd_data(self, team_id: int) -> Dict:
        """
        Fetch the same YTD JSON your dashboard uses, with a proper master session
        and a company in session.
        """
        try:
            from flask import current_app
            from app.models.advisor import Advisor
            from app.config.session import SessionManager
            import calendar

            # --- derive start/end like rest of the service (previous month on the 1st) ---
            today = datetime.now()
            if today.day == 1:
                if today.month == 1:
                    y = today.year - 1
                    m = 12
                else:
                    y = today.year
                    m = today.month - 1
                start_date = datetime(y, m, 1)
                end_date = datetime(y, m, calendar.monthrange(y, m)[1])
            else:
                start_date = datetime(today.year, today.month, 1)
                end_date = today

            # --- make a real request to the controller's route, with a valid session ---
            with current_app.test_client() as client:
                # seed session: pick any master user and set current company
                with client.session_transaction() as sess:
                    master = Advisor.query.filter_by(is_master=True).first()
                    if not master:
                        return {'error': 'No master user found to authorize YTD request'}
                    sess['user_id'] = master.id
                    # set current company for the controller
                    # if your SessionManager has a helper, use it; otherwise set the key used in your app.
                    try:
                        SessionManager.set_current_company(sess, master.company)
                    except Exception:
                        sess['current_company'] = getattr(master, 'company', 'windsor')

                resp = client.get(
                    f"/api/teams/ytd-performance/{team_id}",
                    query_string={
                        "start_date": start_date.strftime("%Y-%m-%d"),
                        "end_date": end_date.strftime("%Y-%m-%d"),
                    },
                )

            # parse JSON and handle non-200s
            if resp.is_json:
                data = resp.get_json()
            else:
                return {'error': f'Unexpected response type: {resp.status}'}

            if resp.status_code != 200:
                # bubble up any error field the controller returned
                return data if isinstance(data, dict) else {'error': f'HTTP {resp.status_code}'}

            return data or {'error': 'Empty JSON from YTD endpoint'}
        except Exception as e:
            logger.error(f"Error getting YTD team data: {e}")
            return {'error': str(e)}

    def _generate_excel_from_ytd_data(self, team_data, team_name, report_month):
        """Generate Excel file from YTD data structure"""
        try:
            import pandas as pd
            import io
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Extract the most recent month's data
            monthly_data = team_data.get('monthly_data', [])
            if not monthly_data:
                raise ValueError("No monthly data available")
            
            # Use the last month's data (most recent)
            latest_month = monthly_data[-1]
            members_data = latest_month.get('members', [])
            
            # Create DataFrame with same structure as your screenshot
            rows = []
            for member in members_data:
                row = {
                    'Advisor': member.get('advisor', ''),
                    'Appointments Booked': member.get('appointments_booked', 0),
                    'Appointments Completed': member.get('appointments_completed', 0),
                    'Outbound Calls': member.get('outbound_calls', 0),
                    'Total Activity': member.get('total_activity', 0),
                    'Number of M Apps': member.get('mortgage_apps', 0),
                    'Insurance Apps': member.get('insurance_apps', 0),
                    'Insurance Referrals': member.get('insurance_referrals', 0),
                    'Other Referrals': member.get('other_referrals', 0),
                    'Conversion %': f"{member.get('conversion_rate', 0)}%",
                    'Submitted': f"£{member.get('submitted_total', 0):,.2f}",
                    'Total': f"£{member.get('submitted_total', 0):,.2f}",
                    'Target': f"£{member.get('monthly_target', 0):,.2f}",
                    'Vs Target': f"£{member.get('vs_target', 0):,.2f}"
                }
                
                # Add C&C Apps column if it's CnC company
                if 'cnc_apps' in member:
                    row['C&C Apps'] = member.get('cnc_apps', 0)
                    
                rows.append(row)
            
            # Add totals row from YTD data
            totals = latest_month.get('totals', {})
            totals_row = {
                'Advisor': 'Totals:',
                'Appointments Booked': totals.get('appointments_booked', 0),
                'Appointments Completed': totals.get('appointments_completed', 0),
                'Outbound Calls': totals.get('outbound_calls', 0),
                'Total Activity': totals.get('total_activity', 0),
                'Number of M Apps': totals.get('mortgage_apps', 0),
                'Insurance Apps': totals.get('insurance_apps', 0),
                'Insurance Referrals': totals.get('insurance_referrals', 0),
                'Other Referrals': totals.get('other_referrals', 0),
                'Conversion %': '-',
                'Submitted': f"£{totals.get('submitted_total', 0):,.2f}",
                'Total': f"£{totals.get('submitted_total', 0):,.2f}",
                'Target': f"£{totals.get('monthly_target', 0):,.2f}",
                'Vs Target': f"£{(totals.get('submitted_total', 0) - totals.get('monthly_target', 0)):,.2f}"
            }
            
            if 'cnc_apps' in totals:
                totals_row['C&C Apps'] = totals.get('cnc_apps', 0)
            
            rows.append(totals_row)
            
            # Create DataFrame
            df = pd.DataFrame(rows)
            
            # Write to Excel
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                sheet_name = f"{report_month} {team_name}"
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                
                # Basic formatting
                wb = writer.book
                ws = wb[sheet_name]
                
                # Header formatting
                header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                center = Alignment(horizontal="center", vertical="center")
                
                # Apply header formatting
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center
                
                # Auto-size columns
                for column in ws.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
                
                # Bold totals row
                last_row = ws.max_row
                for cell in ws[last_row]:
                    cell.font = Font(bold=True)
            
            buf.seek(0)
            return buf.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating Excel from YTD data: {e}")
            # Fallback to basic Excel
            return self._generate_basic_excel(team_name, report_month)
    
    def _generate_basic_excel(self, team_name, report_month):
        """Generate basic Excel file as fallback"""
        try:
            import pandas as pd
            import io
            
            # Create basic structure
            data = {
                'Advisor': ['No Data Available'],
                'Note': ['Please check team configuration and data availability']
            }
            
            df = pd.DataFrame(data)
            buf = io.BytesIO()
            
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=f"{report_month} {team_name}")
            
            buf.seek(0)
            return buf.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating basic Excel: {e}")
            return b"Error generating Excel file"
    
    def _calculate_team_summary_from_ytd(self, team_data):
        """Calculate team summary from YTD data structure"""
        try:
            monthly_data = team_data.get('monthly_data', [])
            if not monthly_data:
                return {
                    'total_submitted': 0,
                    'total_applications': 0,
                    'team_size': 0,
                    'conversion_rate': 0
                }
            
            # Use the latest month's data
            latest_month = monthly_data[-1]
            totals = latest_month.get('totals', {})
            members = latest_month.get('members', [])
            
            # Calculate total applications
            total_apps = (totals.get('mortgage_apps', 0) + 
                         totals.get('insurance_apps', 0) + 
                         totals.get('cnc_apps', 0))
            
            # Calculate average conversion rate
            conversion_rates = [m.get('conversion_rate', 0) for m in members if m.get('conversion_rate', 0) > 0]
            avg_conversion = sum(conversion_rates) / len(conversion_rates) if conversion_rates else 0
            
            return {
                'total_submitted': totals.get('submitted_total', 0),
                'total_applications': total_apps,
                'team_size': len(members),
                'conversion_rate': avg_conversion
            }
            
        except Exception as e:
            logger.error(f"Error calculating team summary from YTD: {e}")
            return {
                'total_submitted': 0,
                'total_applications': 0,
                'team_size': 0,
                'conversion_rate': 0
            }
    
    def _determine_report_month(self):
        """Determine which month to report on - previous month if it's the 1st of current month"""
        today = datetime.now()
        
        if today.day == 1:
            # It's the 1st of the month, use previous month
            if today.month == 1:
                # January 1st, use December of previous year
                report_month = f"{today.year - 1}-12"
                month_display = f"December {today.year - 1}"
            else:
                # Use previous month of current year
                prev_month = today.month - 1
                report_month = f"{today.year}-{prev_month:02d}"
                month_display = f"{datetime(today.year, prev_month, 1).strftime('%B %Y')}"
            
            logger.info(f"First of month detected, using previous month: {month_display}")
        else:
            # Use current month
            report_month = f"{today.year}-{today.month:02d}"
            month_display = f"{today.strftime('%B %Y')}"
        
        return report_month, month_display
    
    def _calculate_team_summary(self, team_data: Dict) -> Dict:
        """Calculate summary metrics for email display - fallback method for old report service"""
        if not team_data or 'rows' not in team_data:
            return {
                'total_submitted': 0,
                'total_applications': 0,
                'team_size': 0,
                'conversion_rate': 0
            }
            
        rows = team_data['rows']
        
        # Filter out totals row if present
        data_rows = [row for row in rows if row.get('Advisor', '').lower() != 'totals']
        
        total_submitted = 0
        total_apps = 0
        conversions = []
        
        for row in data_rows:
            # Handle different possible column names for submitted amounts
            submitted_value = row.get('Total', 0) or row.get('Submitted', 0) or row.get('Total Submitted', 0)
            if isinstance(submitted_value, (int, float)):
                total_submitted += submitted_value
            elif isinstance(submitted_value, str):
                # Remove currency symbols and convert
                try:
                    clean_value = submitted_value.replace('£', '').replace(',', '')
                    total_submitted += float(clean_value)
                except (ValueError, AttributeError):
                    pass
            
            # Handle different possible column names for applications
            apps_value = (row.get('Apps', 0) or row.get('Applications', 0) or 
                         row.get('Total Apps', 0) or row.get('Number of M Apps', 0) or
                         (row.get('Mortgage Apps', 0) + row.get('Insurance Apps', 0) + row.get('C&C Apps', 0)))
            
            if isinstance(apps_value, (int, float)):
                total_apps += apps_value
            
            # Handle conversion rate
            conversion_value = row.get('Conversion %', 0) or row.get('Conversion', 0)
            if isinstance(conversion_value, str):
                try:
                    conversion_value = float(conversion_value.replace('%', ''))
                except (ValueError, AttributeError):
                    conversion_value = 0
            
            if isinstance(conversion_value, (int, float)) and conversion_value > 0:
                conversions.append(conversion_value)
        
        # Calculate average conversion rate
        avg_conversion = sum(conversions) / len(conversions) if conversions else 0
        
        return {
            'total_submitted': total_submitted,
            'total_applications': total_apps,
            'team_size': len(data_rows),
            'conversion_rate': avg_conversion
        }
    
    def _send_via_smtp(self, sender_email: str, recipient_emails: List[str], subject: str, 
                       text_body: str, html_body: str, attachment_bytes: bytes, attachment_filename: str) -> bool:
        """Send email via SMTP with attachment"""
        try:
            # Create message
            msg = MIMEMultipart('alternative')
            msg['From'] = sender_email
            msg['To'] = ', '.join(recipient_emails)
            msg['Subject'] = subject
            
            # Add text and HTML versions
            msg.attach(MIMEText(text_body, 'plain', 'utf-8'))
            msg.attach(MIMEText(html_body, 'html', 'utf-8'))
            
            # Add Excel attachment
            attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            attachment.set_payload(attachment_bytes)
            encoders.encode_base64(attachment)
            attachment.add_header(
                'Content-Disposition',
                f'attachment; filename= {attachment_filename}'
            )
            msg.attach(attachment)
            
            # Create SMTP connection
            if self.use_tls:
                context = ssl.create_default_context()
                server = smtplib.SMTP(self.smtp_server, self.smtp_port)
                server.starttls(context=context)
            else:
                server = smtplib.SMTP_SSL(self.smtp_server, self.smtp_port)
            
            # Login and send
            server.login(self.username, self.password)
            text = msg.as_string()
            server.sendmail(sender_email, recipient_emails, text)
            server.quit()
            
            logger.info(f"Email sent successfully to {len(recipient_emails)} recipients")
            return True
            
        except smtplib.SMTPAuthenticationError as e:
            logger.error(f"SMTP Authentication failed: {e}")
            return False
        except smtplib.SMTPRecipientsRefused as e:
            logger.error(f"SMTP Recipients refused: {e}")
            return False
        except smtplib.SMTPServerDisconnected as e:
            logger.error(f"SMTP Server disconnected: {e}")
            return False
        except Exception as e:
            logger.error(f"SMTP send error: {e}")
            return False
    
    def test_connection(self) -> bool:
        """Test SMTP connection and authentication"""
        try:
            # Create SMTP connection
            if self.use_tls:
                context = ssl.create_default_context()
                server = smtplib.SMTP(self.smtp_server, self.smtp_port)
                server.starttls(context=context)
            else:
                server = smtplib.SMTP_SSL(self.smtp_server, self.smtp_port)
            
            # Test login
            server.login(self.username, self.password)
            server.quit()
            
            logger.info("SMTP connection test successful")
            return True
            
        except smtplib.SMTPAuthenticationError:
            logger.error("SMTP Authentication failed - check username/password")
            return False
        except smtplib.SMTPServerDisconnected:
            logger.error("SMTP Server connection failed - check server/port")
            return False
        except Exception as e:
            logger.error(f"SMTP connection test failed: {e}")
            return False