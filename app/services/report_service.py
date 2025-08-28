from app.models.advisor import Advisor
from app.models.submission import Submission
from app.models.paid_case import PaidCase
from app.services.date import DateService
from app.config.settings import ConfigurationManager
from app.config.session import SessionManager
from app.models.referral_recipient import ReferralRecipient
from app.models import db
from app.controllers.api import APIController
from app.models.referral_mapping import ReferralMapping
from app.models.team import Team
import io
import pandas as pd
from flask import session, request, jsonify, current_app

class ReportService:
    """Service for generating and exporting reports"""
    
    def __init__(self):
        self.date_service = DateService()
        self.config = ConfigurationManager()
        # self.api_controller = APIController(current_app)

    def _month_dates(self):
        start, end = DateService.get_current_month_dates()
        return start, end

    def _normalize(self, s):
        return (s or "").strip().lower()

    def _check_referral_match(self, referral_to_value: str, target_full_name: str, target_id: int, company: str) -> bool:
        """
        Matches a free-text 'referral_to' to a specific advisor using DB mappings
        and simple name checks (like APIController._check_referral_match).
        """
        if not referral_to_value or not target_full_name:
            return False

        ref_l = self._normalize(referral_to_value)
        tgt_l = self._normalize(target_full_name)

        # 1) DB mapping (exact, canonical)
        mapped_id = ReferralMapping.get_advisor_for_referral(ref_l, company)
        if mapped_id == target_id:
            return True

        # 2) Hardcoded/company mappings (if present)
        # Pull the same mapping source your config uses (company name_mappings)
        try:
            from app.config import config_manager
            cc = config_manager.get_company_config(company)
            name_map = getattr(cc, 'name_mappings', {}) or {}
            mapped_name = name_map.get(ref_l)
            if mapped_name and self._normalize(mapped_name) == tgt_l:
                return True
        except Exception:
            pass

        # 3) Loose checks
        if tgt_l in ref_l or ref_l in tgt_l:
            return True
        first = tgt_l.split()[0] if tgt_l else ""
        if first and first in ref_l:
            return True

        return False
    def _loosely_matches(self, text, full_name):
        if not text or not full_name:
            return False
        t = text.lower().strip()
        n = full_name.lower().strip()
        if n in t or t in n:
            return True
        parts = [p for p in n.split() if len(p) > 2]
        return any(p in t for p in parts)
    
    def _month_bounds(self):
        start, end = DateService.get_current_month_dates()
        return start, end

    def _is_mortgage_key(self, k: str) -> bool:
        kl = (k or "").lower()
        return "mortgage" in kl or "residential" in kl

    def _is_insurance_key(self, k: str) -> bool:
        kl = (k or "").lower()
        return "insurance" in kl or "protection" in kl

    def _check_referral_match(self, referral_to_value, advisor_full_name, advisor_id, company):
        """Port of the APIController logic so we don’t depend on the controller."""
        if not referral_to_value or not advisor_full_name:
            return False

        from app.models.referral_mapping import ReferralMapping
        referral_to_lower = referral_to_value.lower().strip()
        advisor_name_lower = advisor_full_name.lower().strip()

        mapped_advisor_id = ReferralMapping.get_advisor_for_referral(referral_to_value, company)
        if mapped_advisor_id == advisor_id:
            return True

        # company config mappings (e.g., mike → michael)
        company_cfg = self.config.get_company_config(company)
        if company_cfg and getattr(company_cfg, "name_mappings", None):
            mapped_name = company_cfg.name_mappings.get(referral_to_lower)
            if mapped_name and mapped_name.lower() == advisor_name_lower:
                return True

        if advisor_name_lower in referral_to_lower or referral_to_lower in advisor_name_lower:
            return True

        first = advisor_name_lower.split()[0] if advisor_name_lower else ""
        return bool(first and first in referral_to_lower)

    # --- inside ReportService class (main builder) ---
    def generate_team_monthly_table(self, team_id: int):
        """
        Build the screenshot-style monthly team report using ONLY team data.
        Columns:
          Advisor | Mortgage Apps | Insurance Apps | Insurance Referral | Other Referrals |
          Conversion | Fees | Submitted | Total | Target | Vs Target
        """
        team = db.session.get(Team, team_id)
        if not team:
            return {'error': 'Team not found'}

        company = SessionManager.get_current_company(session)

        # Current month bounds (inclusive)
        start_date, end_date = DateService.get_current_month_dates()


        rows = []
        totals = {
            'Mortgage Apps': 0, 'Insurance Apps': 0, 'Insurance Referral': 0, 'Other Referrals': 0,
            'Fees': 0.0, 'Submitted': 0.0, 'Total': 0.0, 'Target': 0.0, 'Vs Target': 0.0
        }

        # Preload insurance recipients (so we can split referral types)
        insurance_recipient_ids = {r.advisor_id for r in ReferralRecipient.get_recipients_for_company(company)}
        valid_submission_types = self.config.get_valid_business_types(company) or []
        valid_paid_types = self.config.get_valid_paid_case_types(company) or []

        for advisor in team.members:
            # 1) Use the model's metrics (same path as your dashboard)
            m = advisor.calculate_metrics_for_period(
                company, start_date, end_date,
                valid_submission_types, valid_paid_types
            )  # uses your Advisor model logic :contentReference[oaicite:0]{index=0}

            apps = m.get('applications', {}) or {}
            # Robust app split based on business type labels in your config
            mortgage_apps  = sum(v for k, v in apps.items() if self._is_mortgage_key(k))
            insurance_apps = sum(v for k, v in apps.items() if self._is_insurance_key(k))
            print(apps.items())
            # Referrals: keep your current logic for incoming insurance referrals
            all_subs = advisor.get_submissions_for_period(company, start_date, end_date, None)  # :contentReference[oaicite:1]{index=1}
            referrals = [s for s in all_subs if (s.business_type or "").lower() == "referral"]
            ins_ref, other_ref = 0, 0
            if referrals:
                from app.models.advisor import Advisor as AdvisorModel
                insurance_recipient_ids = {r.advisor_id for r in ReferralRecipient.get_recipients_for_company(company)}  # 
                for r in referrals:
                    matched_insurance = False
                    if r.referral_to and insurance_recipient_ids:
                        for rid in insurance_recipient_ids:
                            rec = db.session.get(AdvisorModel, rid)
                            if rec and self._loosely_matches(r.referral_to, rec.full_name):
                                matched_insurance = True
                                break
                    ins_ref += 1 if matched_insurance else 0
                    other_ref += 0 if matched_insurance else 1

            # 2) Fees & Submitted (proc) from metrics (aligns with dashboard)
            total_fee = float(m.get('total_fee', 0.0))
            total_submitted = float(m.get('total_submitted', 0.0))
            submitted_proc = max(0.0, total_submitted - total_fee)

            total = submitted_proc + total_fee

            # Target from advisor's company goal (team goal supported) :contentReference[oaicite:3]{index=3}
            yearly = float(advisor.get_yearly_goal_for_company(company) or 0.0)
            monthly_target = yearly / 12.0
            vs_target = total - monthly_target

            row = {
                "Advisor": advisor.full_name,
                "Mortgage Apps": int(mortgage_apps),
                "Insurance Apps": int(insurance_apps),
                "Insurance Referral": int(ins_ref),
                "Other Referrals": int(other_ref),
                "Conversion": "",
                "Fees": round(total_fee, 2),
                "Submitted": round(submitted_proc, 2),
                "Total": round(total, 2),
                "Target": round(monthly_target, 2),
                "Vs Target": round(vs_target, 2),
            }
            rows.append(row)

            # accumulate totals
            totals['Mortgage Apps']     += mortgage_apps
            totals['Insurance Apps']    += insurance_apps
            totals['Insurance Referral']+= ins_ref
            totals['Other Referrals']   += other_ref
            totals['Fees']              += total_fee
            totals['Submitted']         += submitted_proc
            totals['Total']             += total
            totals['Target']            += monthly_target
            totals['Vs Target']         += vs_target
        print(totals)
        rows.append({
            "Advisor": "Totals:",
            "Mortgage Apps": totals['Mortgage Apps'],
            "Insurance Apps": totals['Insurance Apps'],
            "Insurance Referral": totals['Insurance Referral'],
            "Other Referrals": totals['Other Referrals'],
            "Conversion": "",
            "Fees": round(totals['Fees'], 2),
            "Submitted": round(totals['Submitted'], 2),
            "Total": round(totals['Total'], 2),
            "Target": round(totals['Target'], 2),
            "Vs Target": round(totals['Vs Target'], 2),
        })
        print(rows)
        return {
            "team": {"id": team.id, "name": team.name, "company": team.company},
            "period": {"start": start_date.isoformat(), "end": end_date.isoformat()},
            "rows": rows
        }
    def generate_report(self):
        """Generate report data for the current user (no controller dependency)"""
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401

        advisor = db.session.get(Advisor, user_id)
        if not advisor:
            return jsonify({'error': 'Advisor not found'}), 404

        current_company = SessionManager.get_current_company(session)
        period = request.args.get('period', 'month')
        start_str = request.args.get('start')
        end_str = request.args.get('end')

        start_date, end_date = DateService.resolve_period_dates(period, start_str, end_str)

        # Use the same underlying model logic your API uses
        metrics = advisor.calculate_metrics_for_period(
            current_company, start_date, end_date,
            self.config.get_valid_business_types(current_company),
            self.config.get_valid_paid_case_types(current_company)
        )

        # If you still need raw rows:
        submissions = Submission.query.filter(
            Submission.advisor_id == advisor.id,
            Submission.company == current_company,
            Submission.submission_date >= start_date,
            Submission.submission_date <= end_date
        ).all()

        paid_cases = PaidCase.query.filter(
            PaidCase.advisor_id == advisor.id,
            PaidCase.company == current_company,
            PaidCase.date_paid >= start_date,
            PaidCase.date_paid <= end_date
        ).all()

        report_data = {
            'advisor_name': advisor.full_name,
            'company': current_company,
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'metrics': metrics,
            'submissions': [self._serialize_submission(s) for s in submissions],
            'paid_cases': [self._serialize_paid_case(pc) for pc in paid_cases],
        }
        return jsonify(report_data)
    def export_team_monthly_excel(self, team_id: int) -> bytes:
        """
        Render the same table to an Excel file (single sheet) with simple styling.
        """
        data = self.generate_team_monthly_table(team_id)
        if isinstance(data, dict) and data.get('error'):
            raise ValueError(data['error'])

        df = pd.DataFrame(data['rows'])

        # Write to xlsx in-memory
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            # Sheet name like "June WHM"
            from datetime import date
            sheet_name = f"{date.today().strftime('%B')} {data['team']['name']}"
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            # Basic formatting
            wb = writer.book
            ws = wb[sheet_name]

            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center = Alignment(horizontal="center", vertical="center")

            # header style
            for c in ws[1]:
                c.fill = header_fill
                c.font = header_font
                c.alignment = center

            # widths
            widths = [20, 14, 14, 18, 16, 12, 12, 14, 12, 12, 14]
            from openpyxl.utils import get_column_letter
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            # currency columns: Fees, Submitted, Total, Target, Vs Target
            money_cols = ("G", "H", "I", "J", "K")
            for col in money_cols:
                for cell in ws[col][1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '£#,##0'

            # totals row bold
            last = ws.max_row
            for c in ws[last]:
                c.font = Font(bold=True)

        buf.seek(0)
        return buf.getvalue()