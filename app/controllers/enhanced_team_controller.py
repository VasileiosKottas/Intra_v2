# app/controllers/enhanced_team_controller.py
"""
Enhanced Team Performance Report Controller with YTD data and Calendly integration
"""

from flask import request, jsonify, session, send_file
from datetime import datetime, timedelta
from app.controllers.base import BaseController
from app.config.session import SessionManager
from app.config.settings import ConfigurationManager
from app.services.calendly_service import CalendlyService
from app.models.team import Team
from app.models.advisor import Advisor
from app.models.submission import Submission
from app.models.paid_case import PaidCase
from app.models import db
import calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # Add Border, Side
from sqlalchemy import and_, or_
import io

def _eod(d: datetime) -> datetime:
    # end-of-day helper (inclusive range end)
    return d.replace(hour=23, minute=59, second=59, microsecond=999999)


class EnhancedTeamReportController(BaseController):
    """Enhanced team reporting with YTD data and real API integration"""
    
    def __init__(self, app):
        super().__init__(app)
        self.config_manager = ConfigurationManager()
        self.calendly_service = CalendlyService()
    
    def register_routes(self):
        """Register enhanced team report routes"""
        print("Registering enhanced team report routes...")
        
        # Team performance report with YTD data
        self.app.add_url_rule('/api/teams/ytd-performance/<int:team_id>', 
                             'api.team_ytd_performance',
                             self.master_required(self.get_ytd_performance_report), 
                             methods=['GET'])
        
        # YTD Totals dashboard data
        self.app.add_url_rule('/api/teams/ytd-totals/<int:team_id>', 
                             'api.team_ytd_totals',
                             self.master_required(self.get_ytd_totals), 
                             methods=['GET'])
        
        # Available teams
        self.app.add_url_rule('/api/teams/performance-available', 
                             'api.teams_performance_available',
                             self.master_required(self.get_available_teams), 
                             methods=['GET'])
        
        # Excel download
        self.app.add_url_rule('/api/teams/ytd-excel/<int:team_id>',
                             'api.team_ytd_excel',
                             self.master_required(self.download_ytd_excel),
                             methods=['GET'])
        
        # Debug routes
        self.app.add_url_rule('/api/test/calendly-debug', 'api.test_calendly_debug',
                             self.test_calendly_debug, methods=['GET'])
        
        self.app.add_url_rule('/api/test/calendly-emails', 'api.test_calendly_emails',
                             self.test_calendly_emails, methods=['GET'])
        
        # Pipeline placeholder
        self.app.add_url_rule('/api/teams/pipeline-summary/<int:team_id>', 
                             'api.team_pipeline_summary',
                             self.master_required(self.get_pipeline_summary), 
                             methods=['GET'])

        self.app.add_url_rule('/api/test/excel-download',
                            'api.test_excel_download',
                            self.master_required(self.test_excel_download),
                            methods=['GET'])
        

        print("Enhanced team report routes registered successfully")
    
    def get_ytd_performance_report(self, team_id):
        """Generate YTD performance report with monthly breakdown and real Calendly data"""
        try:
            user = self.get_current_user()
            if not user or not user.is_master:
                return jsonify({'error': 'Access denied'}), 403
            
            current_company = SessionManager.get_current_company(session)
            team = Team.query.filter_by(id=team_id, company=current_company).first()
            if not team:
                return jsonify({'error': 'Team not found'}), 404
            
            # Parse date parameters
            end_date_str = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
            start_date_str = request.args.get('start_date')
            
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            if start_date_str:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            else:
                start_date = datetime(end_date.year, 1, 1)
            
            print(f"Getting YTD data for team {team.name} from {start_date} to {end_date}")
            
            # Generate monthly data
            monthly_data = []
            current_month = start_date.replace(day=1)
            
            while current_month <= end_date:
                month_end = self._get_month_end(current_month, end_date)
                month_start = current_month
                month_name = current_month.strftime('%B %Y')
                
                print(f"Processing month: {month_name}")
                
                month_members = []
                
                for member in team.members:
                    print(f"Processing member: {member.full_name}")
                    
                    # Calculate submission metrics for current company only
                    submission_metrics = member.calculate_metrics_for_period(
                        current_company, month_start, month_end,
                        self.config_manager.get_valid_business_types(current_company),
                        self.config_manager.get_valid_paid_case_types(current_company)
                    )
                    
                    # USE THE EXISTING total_submitted calculation
                    submitted_total = submission_metrics.get('total_submitted', 0)
                    monthly_target = self._calculate_monthly_target(member, month_start, current_company)
                    
                    # Get real Calendly data for this member
                    appointments_booked, appointments_completed = self._get_real_calendly_data(
                        member, month_start, month_end
                    )
                    
                    # Placeholder for ALTOS calls (until you get the token)
                    outbound_calls = 85  
                    total_activity = outbound_calls + appointments_completed
                    
                    # Company-specific applications
                    apps_data = self._get_company_specific_apps(member, month_start, month_end, current_company)
                    
                    # Calculate conversion
                    total_apps = apps_data['total_apps']
                    conversion_rate = (total_apps / appointments_completed * 100) if appointments_completed > 0 else 0
                    
                    member_data = {
                        'advisor': member.full_name,
                        'appointments_booked': appointments_booked,
                        'appointments_completed': appointments_completed,
                        'outbound_calls': outbound_calls,
                        'total_activity': total_activity,
                        'mortgage_apps': apps_data['mortgage_apps'],
                        'insurance_apps': apps_data['insurance_apps'],
                        'cnc_apps': apps_data['cnc_apps'],
                        'insurance_referrals': apps_data['insurance_referrals'],
                        'other_referrals': apps_data['other_referrals'],
                        'submitted_total': submitted_total,
                        'conversion_rate': round(conversion_rate, 1),
                        'monthly_target': monthly_target,
                        'vs_target': submitted_total - monthly_target
                    }
                    
                    month_members.append(member_data)
                
                monthly_data.append({
                    'month': month_name,
                    'month_key': current_month.strftime('%Y-%m'),
                    'members': month_members,
                    'totals': self._calculate_month_totals(month_members, current_company)
                })
                
                # Move to next month
                if current_month.month == 12:
                    current_month = current_month.replace(year=current_month.year + 1, month=1)
                else:
                    current_month = current_month.replace(month=current_month.month + 1)
                
                if current_month > end_date:
                    break
            
            return jsonify({
                'success': True,
                'team_name': team.name,
                'team_id': team.id,
                'company': current_company,
                'date_range': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat(),
                    'period': f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
                },
                'monthly_data': monthly_data,
                'ytd_totals': self._calculate_ytd_totals(monthly_data)
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500
    
    def _get_member_email(self, member):
        """Get member email with fallback"""
        if hasattr(member, 'email') and member.email:
            return member.email.lower()
        
        # Generate email from name as fallback  
        return f"{member.full_name.lower().replace(' ', '.')}@company.com"
        
    def _get_company_specific_apps(self, member, start_date, end_date, company):
            """Get application data specific to the current company"""
            apps_data = {
                'mortgage_apps': 0,
                'insurance_apps': 0, 
                'cnc_apps': 0,
                'insurance_referrals': 0,
                'other_referrals': 0,
                'total_apps': 0
            }
            
            print(member.calculate_metrics_for_period(company, start_date, end_date, \
                                                    self.config_manager.get_valid_business_types(company),
                                                    self.config_manager.get_valid_paid_case_types(company)))
            
            apps = member.calculate_metrics_for_period(company, start_date, end_date, \
                                                    self.config_manager.get_valid_business_types(company),
                                                    self.config_manager.get_valid_paid_case_types(company))

            applications = apps.get('applications', {})
            
            # Calculate mortgage apps - EXCLUDE Personal Insurance (Including GI)
            mortgage_apps = 0
            insurance_apps = 0
            
            for app_type, count in applications.items():
                if app_type == 'Personal Insurance (Including GI)':
                    # Personal Insurance goes to insurance_apps
                    insurance_apps += count
                else:
                    # Everything else goes to mortgage_apps (Residential Mortgage, Product Transfer, etc.)
                    mortgage_apps += count
            
            print(f"Applications breakdown for {member.full_name}:")
            print(f"  Raw applications: {applications}")
            print(f"  Mortgage apps (excluding Personal Insurance): {mortgage_apps}")
            print(f"  Insurance apps (Personal Insurance only): {insurance_apps}")
            
            # Company-specific business type filtering
            if company.lower() == 'windsor':
                # Windsor: Mortgage and Insurance only, NO C&C
                apps_data['mortgage_apps'] = mortgage_apps
                apps_data['insurance_apps'] = insurance_apps
                apps_data['cnc_apps'] = 0  # Explicitly set to 0 for Windsor
                
            elif company.lower() == 'cnc':
                # C&C: All types including Conveyancing
                apps_data['mortgage_apps'] = mortgage_apps
                apps_data['insurance_apps'] = insurance_apps
                apps_data['cnc_apps'] = 0  # You can adjust this if C&C has specific app types

            # Get referral submissions by this advisor
            referral_submissions = Submission.query.filter(
                and_(
                    or_(
                        Submission.advisor_id == member.id,
                        and_(Submission.advisor_id.is_(None), Submission.advisor_name == member.full_name)
                    ),
                    Submission.company == company,
                    Submission.submission_date >= start_date,
                    Submission.submission_date <= end_date,
                    Submission.business_type == 'Referral'
                )
            ).all()
            
            print(f"Found {len(referral_submissions)} referral submissions BY {member.full_name}")

            # Get all advisors in the database to check against
            all_advisors = Advisor.query.all()
            advisor_names = set()
            for advisor in all_advisors:
                # Add full name and common variations
                advisor_names.add(advisor.full_name.lower().strip())
                # Add first name only
                first_name = advisor.full_name.split()[0].lower().strip()
                advisor_names.add(first_name)
                # Add any email-based names if they exist
                if hasattr(advisor, 'email') and advisor.email:
                    email_name = advisor.email.split('@')[0].lower().strip()
                    advisor_names.add(email_name)
            
            print(f"Known advisor names: {sorted(advisor_names)}")

            insurance_referrals = 0  # Referrals TO known advisors in database
            other_referrals = 0      # Survey referral, Conveyancing referral, referrals to external people
            
            for referral in referral_submissions:
                referral_to = (referral.referral_to or '').lower().strip()
                original_type = getattr(referral, 'original_business_type', '') or ''
                
                print(f"  Processing: '{original_type}' -> referral_to: '{referral_to}'")
                
                # Check if referral_to matches any known advisor
                is_to_advisor = False
                if referral_to:
                    # Check exact match
                    if referral_to in advisor_names:
                        is_to_advisor = True
                    else:
                        # Check if referral_to contains any advisor name or vice versa
                        for advisor_name in advisor_names:
                            if advisor_name in referral_to or referral_to in advisor_name:
                                is_to_advisor = True
                                break
                
                if is_to_advisor:
                    insurance_referrals += 1
                    print(f"    -> Insurance Referral (to advisor: {referral_to})")
                else:
                    other_referrals += 1
                    print(f"    -> Other Referral (external or survey/conveyancing: {referral_to or original_type})")
            
            print(f"  Insurance referrals (to known advisors): {insurance_referrals}")
            print(f"  Other referrals (external/survey/conveyancing): {other_referrals}")
            
            apps_data['insurance_referrals'] = insurance_referrals
            apps_data['other_referrals'] = other_referrals
            
            # Calculate total apps (excluding referrals since they're counted separately)
            apps_data['total_apps'] = (apps_data['mortgage_apps'] + 
                                    apps_data['insurance_apps'] + 
                                    apps_data['cnc_apps'])
            
            return apps_data
 
    def _calculate_month_totals(self, month_members, company):
        """Calculate monthly totals with company-specific logic"""
        totals = {
            'appointments_booked': sum(m['appointments_booked'] for m in month_members),
            'appointments_completed': sum(m['appointments_completed'] for m in month_members),
            'outbound_calls': sum(m['outbound_calls'] for m in month_members),
            'total_activity': sum(m['total_activity'] for m in month_members),
            'mortgage_apps': sum(m['mortgage_apps'] for m in month_members),
            'insurance_apps': sum(m['insurance_apps'] for m in month_members),
            'insurance_referrals': sum(m['insurance_referrals'] for m in month_members),
            'other_referrals': sum(m['other_referrals'] for m in month_members),
            'submitted_total': sum(m['submitted_total'] for m in month_members),
            'monthly_target': sum(m['monthly_target'] for m in month_members)
        }
        
        # Company-specific totals
        if company.lower() == 'cnc':
            totals['cnc_apps'] = sum(m['cnc_apps'] for m in month_members)
        else:
            totals['cnc_apps'] = 0  # Always 0 for Windsor
        
        return totals
    
    def _calculate_ytd_totals(self, monthly_data):
        """Calculate YTD totals across all months"""
        ytd_totals = {
            'total_activity': 0,
            'submitted_total': 0,
            'appointments_completed': 0,
            'outbound_calls': 0,
            'mortgage_apps': 0,
            'insurance_apps': 0,
            'cnc_apps': 0,
            'insurance_referrals': 0,
            'other_referrals': 0
        }
        
        for month in monthly_data:
            month_totals = month.get('totals', {})
            for key in ytd_totals.keys():
                ytd_totals[key] += month_totals.get(key, 0)
        
        return ytd_totals
    
    def _get_month_end(self, month_start, overall_end):
        """Get the last day of the month, but not beyond overall_end"""
        if month_start.month == 12:
            next_month = month_start.replace(year=month_start.year + 1, month=1)
        else:
            next_month = month_start.replace(month=month_start.month + 1)
        
        month_end = next_month - timedelta(days=1)
        return min(month_end, overall_end)
    
    def _calculate_monthly_target(self, member, month_date, company):
        """Calculate monthly target for member"""
        yearly_goal = member.get_yearly_goal_for_company(company)
        return yearly_goal / 12 if yearly_goal > 0 else 0
    
    def _count_business_type(self, advisor, start_date, end_date, company, business_types):
        """Count submissions by business type for specific company"""
        if isinstance(business_types, str):
            business_types = [business_types]
        
        count = 0
        for business_type in business_types:
            count += Submission.query.filter(
                and_(
                    or_(
                        Submission.advisor_id == advisor.id,
                        and_(Submission.advisor_id.is_(None), Submission.advisor_name == advisor.full_name)
                    ),
                    Submission.company == company,
                    Submission.submission_date >= start_date,
                    Submission.submission_date <= end_date,
                    Submission.business_type.ilike(f'%{business_type}%')
                )
            ).count()
        
        return count
    
    def _count_referrals(self, advisor, start_date, end_date, company, referral_type):
        """Count referrals by type for specific company"""
        base_query = Submission.query.filter(
            and_(
                or_(
                    Submission.advisor_id == advisor.id,
                    and_(Submission.advisor_id.is_(None), Submission.advisor_name == advisor.full_name)
                ),
                Submission.company == company,
                Submission.submission_date >= start_date,
                Submission.submission_date <= end_date,
                Submission.business_type == 'Referral'
            )
        )
        
        if referral_type == 'insurance':
            return base_query.filter(Submission.referral_to.ilike('%insurance%')).count()
        else:
            return base_query.filter(~Submission.referral_to.ilike('%insurance%')).count()
    
    def get_ytd_totals(self, team_id):
        """Get YTD totals for dashboard"""
        try:
            user = self.get_current_user()
            if not user or not user.is_master:
                return jsonify({'error': 'Access denied'}), 403
            
            current_company = SessionManager.get_current_company(session)
            team = Team.query.filter_by(id=team_id, company=current_company).first()
            if not team:
                return jsonify({'error': 'Team not found'}), 404
            
            # Parse date parameters
            end_date_str = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
            start_date_str = request.args.get('start_date')
            
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            if start_date_str:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            else:
                start_date = datetime(end_date.year, 1, 1)
            
            # Get YTD data for each member
            ytd_data = {
                'total_submitted': [],
                'appointments_completed': [],
                'total_apps': []
            }
            year = end_date.year
            year_start = datetime(year, 1, 1)
            for member in team.members:
                # Calculate YTD submitted (for the selected period)
                submission_metrics = member.calculate_metrics_for_period(
                    current_company, start_date, end_date,
                    self.config_manager.get_valid_business_types(current_company),
                    self.config_manager.get_valid_paid_case_types(current_company)
                )
                
                # YTD Total Submitted
                ytd_submitted = submission_metrics.get('total_submitted', 0)
                yearly_target = member.get_yearly_goal_for_company(current_company)
                vs_yearly_target = ytd_submitted - yearly_target
                
                # FIXED: Calculate quarters within the user's selected date range

                q1_start, q1_end = datetime(year, 1, 1), min(_eod(datetime(year, 3, 31)), _eod(end_date))
                q2_start, q2_end = datetime(year, 4, 1), min(_eod(datetime(year, 6, 30)), _eod(end_date))
                q3_start, q3_end = datetime(year, 7, 1), min(_eod(datetime(year, 9, 30)), _eod(end_date))
                q4_start, q4_end = datetime(year,10, 1), min(_eod(datetime(year,12, 31)), _eod(end_date))
                # Q1 data (only if the quarter overlaps with selected period)
                # Q1
                if q1_start <= q1_end:
                    q1_submitted_metrics = member.calculate_metrics_for_period(
                        current_company, q1_start, q1_end,
                        self.config_manager.get_valid_business_types(current_company),
                        self.config_manager.get_valid_paid_case_types(current_company)
                    )
                    q1_submitted = q1_submitted_metrics.get('total_submitted', 0)
                    q1_appointments = (self._get_completed_appointments_chunked(member, q1_start, q1_end)
                                    if hasattr(self, '_get_completed_appointments_chunked')
                                    else self._get_real_calendly_appointments(member, q1_start, q1_end))
                    q1_apps = self._get_company_specific_apps(member, q1_start, q1_end, current_company)['total_apps']
                else:
                    q1_submitted = q1_appointments = q1_apps = 0

                # Q2
                if q2_start <= q2_end:
                    q2_submitted_metrics = member.calculate_metrics_for_period(
                        current_company, q2_start, q2_end,
                        self.config_manager.get_valid_business_types(current_company),
                        self.config_manager.get_valid_paid_case_types(current_company)
                    )
                    q2_submitted = q2_submitted_metrics.get('total_submitted', 0)
                    q2_appointments = (self._get_completed_appointments_chunked(member, q2_start, q2_end)
                                    if hasattr(self, '_get_completed_appointments_chunked')
                                    else self._get_real_calendly_appointments(member, q2_start, q2_end))
                    q2_apps = self._get_company_specific_apps(member, q2_start, q2_end, current_company)['total_apps']
                else:
                    q2_submitted = q2_appointments = q2_apps = 0

                # Q3
                if q3_start <= q3_end:
                    q3_submitted_metrics = member.calculate_metrics_for_period(
                        current_company, q3_start, q3_end,
                        self.config_manager.get_valid_business_types(current_company),
                        self.config_manager.get_valid_paid_case_types(current_company)
                    )
                    q3_submitted = q3_submitted_metrics.get('total_submitted', 0)
                    q3_appointments = (self._get_completed_appointments_chunked(member, q3_start, q3_end)
                                    if hasattr(self, '_get_completed_appointments_chunked')
                                    else self._get_real_calendly_appointments(member, q3_start, q3_end))
                    q3_apps = self._get_company_specific_apps(member, q3_start, q3_end, current_company)['total_apps']
                else:
                    q3_submitted = q3_appointments = q3_apps = 0

                # Q4
                if q4_start <= q4_end:
                    q4_submitted_metrics = member.calculate_metrics_for_period(
                        current_company, q4_start, q4_end,
                        self.config_manager.get_valid_business_types(current_company),
                        self.config_manager.get_valid_paid_case_types(current_company)
                    )
                    q4_submitted = q4_submitted_metrics.get('total_submitted', 0)
                    q4_appointments = (self._get_completed_appointments_chunked(member, q4_start, q4_end)
                                    if hasattr(self, '_get_completed_appointments_chunked')
                                    else self._get_real_calendly_appointments(member, q4_start, q4_end))
                    q4_apps = self._get_company_specific_apps(member, q4_start, q4_end, current_company)['total_apps']
                else:
                    q4_submitted = q4_appointments = q4_apps = 0

                # Keep start_date param logic as-is (used for submitted + quarters)
                if start_date_str:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                else:
                    start_date = datetime(end_date.year, 1, 1)

                # --- NEW: define a strict YTD window for apps/appointments ---
                ytd_start = datetime(end_date.year, 1, 1)
                # make the end inclusive (end of day)
                ytd_end = end_date + timedelta(days=1) - timedelta(microseconds=1)
                # YTD appointments and apps (for the full selected period)
                ytd_appointments = self._get_completed_appointments_chunked(member, ytd_start, ytd_end)
                ytd_apps = self._get_company_specific_apps(member, ytd_start, ytd_end, current_company)['total_apps']
                print('ytd app:', ytd_appointments)
                # Total Submitted data
                # Total Submitted data - ADD Q3 and Q4
                ytd_data['total_submitted'].append({
                    'advisor': member.full_name,
                    'q1': q1_submitted,
                    'q2': q2_submitted,
                    'q3': q3_submitted,  # ADD THIS
                    'q4': q4_submitted,  # ADD THIS
                    'ytd_total': ytd_submitted,
                    'yearly_target': yearly_target,
                    'vs_target': vs_yearly_target
                })

                # Appointments Completed data - ADD Q3 and Q4
                ytd_data['appointments_completed'].append({
                    'advisor': member.full_name,
                    'q1': q1_appointments,
                    'q2': q2_appointments,
                    'q3': q3_appointments,  # ADD THIS
                    'q4': q4_appointments,  # ADD THIS
                    'ytd_total': ytd_appointments
                })

                # Total Apps data - ADD Q3 and Q4
                ytd_data['total_apps'].append({
                    'advisor': member.full_name,
                    'q1': q1_apps,
                    'q2': q2_apps,
                    'q3': q3_apps,  # ADD THIS
                    'q4': q4_apps,  # ADD THIS
                    'ytd_total': ytd_apps
                })
        
            return jsonify({
                'success': True,
                'team_name': team.name,
                'company': current_company,
                'date_range': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                },
                'ytd_data': ytd_data
            })
            
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500


    def _get_completed_appointments_chunked(self, member, start_date, end_date):
        """Sum Calendly completed appts across smaller windows to avoid pagination caps."""
        total = 0
        # start at the first of the month for neat month chunks
        cur = start_date.replace(day=1)
        while cur <= end_date:
            # month end, but don’t go past the overall end_date
            if cur.month == 12:
                next_month = cur.replace(year=cur.year + 1, month=1, day=1)
            else:
                next_month = cur.replace(month=cur.month + 1, day=1)
            month_end = min(next_month - timedelta(seconds=1), end_date)
            total += self._get_real_calendly_appointments(member, cur, month_end)
            # next month
            cur = next_month
        return total

    def _get_real_calendly_appointments(self, member, start_date, end_date):
        """Helper method to get Calendly appointments completed for a date range"""
        
        _, appointments_completed = self._get_real_calendly_data(member, start_date, end_date)
        
        return appointments_completed

    
    def _get_quarterly_appointments(self, advisor, company, year, quarter):
        """Get quarterly appointments completed using Calendly data"""
        if quarter == 1:
            start = datetime(year, 1, 1)
            end = datetime(year, 3, 31)
        elif quarter == 2:
            start = datetime(year, 4, 1)
            end = datetime(year, 6, 30)
        elif quarter == 3:
            start = datetime(year, 7, 1)
            end = datetime(year, 9, 30)
        else:  # Q4
            start = datetime(year, 10, 1)
            end = datetime(year, 12, 31)
        
        try:
            # Get Calendly appointments for this quarter
            _, appointments_completed = self._get_real_calendly_data(advisor, start, end)
            return appointments_completed
        except:
            # Fallback to estimated value
            return 25 + (advisor.id % 10)

    def _get_quarterly_apps(self, advisor, company, year, quarter):
        """Get quarterly total apps using existing calculation"""
        if quarter == 1:
            start = datetime(year, 1, 1)
            end = datetime(year, 3, 31)
        elif quarter == 2:
            start = datetime(year, 4, 1)
            end = datetime(year, 6, 30)
        elif quarter == 3:
            start = datetime(year, 7, 1)
            end = datetime(year, 9, 30)
        else:  # Q4
            start = datetime(year, 10, 1)
            end = datetime(year, 12, 31)
        
        # Use the same method as the main dashboard
        apps_data = self._get_company_specific_apps(advisor, start, end, company)
        return apps_data['total_apps']

    def _get_ytd_appointments(self, advisor, company, start_date, end_date):
        """Get YTD appointments completed"""
        try:
            _, appointments_completed = self._get_real_calendly_data(advisor, start_date, end_date)
            return appointments_completed
        except:
            # Fallback to estimated value
            return 60 + (advisor.id % 20)

    def _get_ytd_apps(self, advisor, company, start_date, end_date):
        """Get YTD total apps"""
        apps_data = self._get_company_specific_apps(advisor, start_date, end_date, company)
        return apps_data['total_apps']
    
    def _get_quarterly_submitted(self, advisor, company, year, quarter):
        """Get quarterly submitted amount using existing calculation"""
        if quarter == 1:
            start = datetime(year, 1, 1)
            end = datetime(year, 3, 31)
        elif quarter == 2:
            start = datetime(year, 4, 1)
            end = datetime(year, 6, 30)
        elif quarter == 3:
            start = datetime(year, 7, 1)
            end = datetime(year, 9, 30)
        else:  # Q4
            start = datetime(year, 10, 1)
            end = datetime(year, 12, 31)
        
        metrics = advisor.calculate_metrics_for_period(
            company, start, end,
            self.config_manager.get_valid_business_types(company),
            self.config_manager.get_valid_paid_case_types(company)
        )
        
        return metrics.get('total_submitted', 0)
    
    def get_available_teams(self):
        """Get available teams for dropdown"""
        try:
            user = self.get_current_user()
            if not user or not user.is_master:
                return jsonify({'error': 'Access denied'}), 403
            
            current_company = SessionManager.get_current_company(session)
            teams = Team.query.filter_by(company=current_company).all()
            
            team_data = []
            for team in teams:
                team_data.append({
                    'id': team.id,
                    'name': team.name,
                    'member_count': len(team.members)
                })
            
            return jsonify({
                'success': True,
                'teams': team_data,
                'company': current_company
            })
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500
        
    def download_ytd_excel(self, team_id):
            """Download Excel file for YTD data with 2 sheets: Monthly Performance and YTD Totals"""
            try:
                user = self.get_current_user()
                if not user or not user.is_master:
                    return jsonify({'error': 'Access denied'}), 403
                
                current_company = SessionManager.get_current_company(session)
                team = Team.query.filter_by(id=team_id, company=current_company).first()
                if not team:
                    return jsonify({'error': 'Team not found'}), 404
                
                # Parse date parameters
                end_date_str = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
                start_date_str = request.args.get('start_date')
                
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
                if start_date_str:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                else:
                    start_date = datetime(end_date.year, 1, 1)
                
                # FIXED: Get data directly instead of calling API endpoints
                # Generate performance data directly
                print(f"Getting YTD data for team {team.name} from {start_date} to {end_date}")
                
                # Generate monthly data (same logic as get_ytd_performance_report but return data directly)
                monthly_data = []
                current_month = start_date.replace(day=1)
                
                while current_month <= end_date:
                    month_end = self._get_month_end(current_month, end_date)
                    month_start = current_month
                    month_name = current_month.strftime('%B %Y')
                    
                    print(f"Processing month: {month_name}")
                    
                    month_members = []
                    
                    for member in team.members:
                        # Calculate submission metrics for current company only
                        submission_metrics = member.calculate_metrics_for_period(
                            current_company, month_start, month_end,
                            self.config_manager.get_valid_business_types(current_company),
                            self.config_manager.get_valid_paid_case_types(current_company)
                        )
                        
                        submitted_total = submission_metrics.get('total_submitted', 0)
                        monthly_target = self._calculate_monthly_target(member, month_start, current_company)
                        
                        # Get real Calendly data for this member
                        appointments_booked, appointments_completed = self._get_real_calendly_data(
                            member, month_start, month_end
                        )
                        
                        # Placeholder for ALTOS calls
                        outbound_calls = 85  
                        total_activity = outbound_calls + appointments_completed
                        
                        # Company-specific applications
                        apps_data = self._get_company_specific_apps(member, month_start, month_end, current_company)
                        
                        # Calculate conversion
                        total_apps = apps_data['total_apps']
                        conversion_rate = (total_apps / appointments_completed * 100) if appointments_completed > 0 else 0
                        
                        member_data = {
                            'advisor': member.full_name,
                            'appointments_booked': appointments_booked,
                            'appointments_completed': appointments_completed,
                            'outbound_calls': outbound_calls,
                            'total_activity': total_activity,
                            'mortgage_apps': apps_data['mortgage_apps'],
                            'insurance_apps': apps_data['insurance_apps'],
                            'cnc_apps': apps_data['cnc_apps'],
                            'insurance_referrals': apps_data['insurance_referrals'],
                            'other_referrals': apps_data['other_referrals'],
                            'submitted_total': submitted_total,
                            'conversion_rate': round(conversion_rate, 1),
                            'monthly_target': monthly_target,
                            'vs_target': submitted_total - monthly_target
                        }
                        
                        month_members.append(member_data)
                    
                    monthly_data.append({
                        'month': month_name,
                        'month_key': current_month.strftime('%Y-%m'),
                        'members': month_members,
                        'totals': self._calculate_month_totals(month_members, current_company)
                    })
                    
                    # Move to next month
                    if current_month.month == 12:
                        current_month = current_month.replace(year=current_month.year + 1, month=1)
                    else:
                        current_month = current_month.replace(month=current_month.month + 1)
                    
                    if current_month > end_date:
                        break
                
                # Prepare performance data
                performance_data = {
                    'team_name': team.name,
                    'date_range': {
                        'period': f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
                    },
                    'monthly_data': monthly_data
                }
                
                # Generate YTD totals data directly
                ytd_data = {
                    'total_submitted': [],
                    'appointments_completed': [],
                    'total_apps': []
                }
                
                for member in team.members:
                    # Get submission data using existing method
                    submission_metrics = member.calculate_metrics_for_period(
                        current_company, start_date, end_date,
                        self.config_manager.get_valid_business_types(current_company),
                        self.config_manager.get_valid_paid_case_types(current_company)
                    )
                    
                    # Total Submitted for period
                    period_submitted = submission_metrics.get('total_submitted', 0)
                    yearly_target = member.get_yearly_goal_for_company(current_company)
                    
                    # Calculate vs yearly target
                    vs_yearly_target = period_submitted - yearly_target
                    
                    # Get quarterly submitted data
                    q1_submitted = self._get_quarterly_submitted(member, current_company, start_date.year, 1)
                    q2_submitted = self._get_quarterly_submitted(member, current_company, start_date.year, 2)
                    q3_submitted = self._get_quarterly_submitted(member, current_company, start_date.year, 3)
                    q4_submitted = self._get_quarterly_submitted(member, current_company, start_date.year, 4)

                    # Get quarterly appointments data
                    q1_appointments = self._get_quarterly_appointments(member, current_company, start_date.year, 1)
                    q2_appointments = self._get_quarterly_appointments(member, current_company, start_date.year, 2)
                    q3_appointments = self._get_quarterly_appointments(member, current_company, start_date.year, 3)
                    q4_appointments = self._get_quarterly_appointments(member, current_company, start_date.year, 4)
                    ytd_appointments = self._get_ytd_appointments(member, current_company, start_date, end_date)

                    # Get quarterly apps data
                    q1_apps = self._get_quarterly_apps(member, current_company, start_date.year, 1)
                    q2_apps = self._get_quarterly_apps(member, current_company, start_date.year, 2)
                    q3_apps = self._get_quarterly_apps(member, current_company, start_date.year, 3)
                    q4_apps = self._get_quarterly_apps(member, current_company, start_date.year, 4)
                    ytd_apps = self._get_ytd_apps(member, current_company, start_date, end_date)

                    # Total Submitted data
                    ytd_data['total_submitted'].append({
                        'advisor': member.full_name,
                        'q1': q1_submitted,
                        'q2': q2_submitted,
                        'q3': q3_submitted,
                        'q4': q4_submitted,
                        'ytd_total': period_submitted,
                        'yearly_target': yearly_target,
                        'vs_target': vs_yearly_target
                    })
                    
                    # Appointments Completed data
                    ytd_data['appointments_completed'].append({
                        'advisor': member.full_name,
                        'q1': q1_appointments,
                        'q2': q2_appointments,
                        'q3': q3_appointments,
                        'q4': q4_appointments,
                        'ytd_total': ytd_appointments
                    })

                    # Total Apps data
                    ytd_data['total_apps'].append({
                        'advisor': member.full_name,
                        'q1': q1_apps,
                        'q2': q2_apps,
                        'q3': q3_apps,
                        'q4': q4_apps,
                        'ytd_total': ytd_apps
                    })
                
                # Prepare totals data
                totals_data = {
                    'team_name': team.name,
                    'date_range': {
                        'start': start_date.isoformat(),
                        'end': end_date.isoformat()
                    },
                    'ytd_data': ytd_data
                }
                
                # Create Excel workbook
                wb = openpyxl.Workbook()
                
                # Remove default sheet
                wb.remove(wb.active)
                
                # Create Sheet 1: Monthly Performance
                self._create_monthly_performance_sheet(wb, performance_data, current_company)
                
                # Create Sheet 2: YTD Totals
                self._create_ytd_totals_sheet(wb, totals_data, current_company)
                
                # Save to memory buffer
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                # Create filename
                filename = f"{team.name}_YTD_Report_{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}.xlsx"
                
                return send_file(
                    buffer,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
            except Exception as e:
                import traceback
                traceback.print_exc()
                return jsonify({'error': str(e)}), 500

    def _create_monthly_performance_sheet(self, workbook, data, company):
        """Create the Monthly Performance sheet"""
        ws = workbook.create_sheet("Monthly Performance")
        
        # Styles
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
        title_font = Font(bold=True, size=14)
        totals_font = Font(bold=True, size=10)
        center_alignment = Alignment(horizontal='center', vertical='center')
        currency_alignment = Alignment(horizontal='right', vertical='center')
        
        # Title and metadata
        ws.merge_cells('A1:P1')
        ws['A1'] = f"{data.get('team_name', '')} - Monthly Performance Report"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_alignment
        
        ws['A2'] = f"Company: {company.upper()}"
        ws['A3'] = f"Period: {data.get('date_range', {}).get('period', '')}"
        ws['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        current_row = 6
        
        # Process each month's data
        monthly_data = data.get('monthly_data', [])
        
        for month_data in monthly_data:
            month_name = month_data.get('month', '')
            members = month_data.get('members', [])
            totals = month_data.get('totals', {})
            
            # Month header
            ws.merge_cells(f'A{current_row}:P{current_row}')
            ws[f'A{current_row}'] = month_name
            ws[f'A{current_row}'].font = Font(bold=True, size=12)
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
            
            # Column headers
            headers = [
                'Advisor', 'Appointments\nBooked', 'Appointments\nCompleted', 
                'Outbound\nCalls', 'Total\nActivity', 'Mortgage\nApps',
                'Insurance\nApps'
            ]
            
            # Add C&C Apps column only for CNC company
            if company.lower() == 'cnc':
                headers.append('C&C\nApps')
            
            headers.extend([
                'Insurance\nReferrals', 'Other\nReferrals', 'Submitted\n(Plus Fees)',
                'Conversion\n%', 'Monthly\nTarget', 'Vs\nTarget'
            ])
            
            # Write headers
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            current_row += 1
            
            # Write member data
            for member in members:
                row_data = [
                    member.get('advisor', ''),
                    member.get('appointments_booked', 0),
                    member.get('appointments_completed', 0),
                    member.get('outbound_calls', 0),
                    member.get('total_activity', 0),
                    member.get('mortgage_apps', 0),
                    member.get('insurance_apps', 0)
                ]
                
                # Add C&C Apps only for CNC
                if company.lower() == 'cnc':
                    row_data.append(member.get('cnc_apps', 0))
                
                row_data.extend([
                    member.get('insurance_referrals', 0),
                    member.get('other_referrals', 0),
                    member.get('submitted_total', 0),
                    f"{member.get('conversion_rate', 0)}%",
                    member.get('monthly_target', 0),
                    member.get('vs_target', 0)
                ])
                
                for col, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    
                    # Apply currency formatting to money columns
                    if col in [11, 13, 14]:  # Submitted, Target, Vs Target
                        if isinstance(value, (int, float)):
                            cell.number_format = '£#,##0'
                            cell.alignment = currency_alignment
                    
                current_row += 1
            
            # Totals row
            totals_row = [
                'TOTALS:',
                totals.get('appointments_booked', 0),
                totals.get('appointments_completed', 0),
                totals.get('outbound_calls', 0),
                totals.get('total_activity', 0),
                totals.get('mortgage_apps', 0),
                totals.get('insurance_apps', 0)
            ]
            
            if company.lower() == 'cnc':
                totals_row.append(totals.get('cnc_apps', 0))
            
            totals_row.extend([
                totals.get('insurance_referrals', 0),
                totals.get('other_referrals', 0),
                totals.get('submitted_total', 0),
                '-',  # No conversion rate for totals
                totals.get('monthly_target', 0),
                totals.get('submitted_total', 0) - totals.get('monthly_target', 0)  # Vs Target
            ])
            
            for col, value in enumerate(totals_row, start=1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = totals_font
                
                # Apply currency formatting to money columns
                if col in [11, 13, 14]:  # Submitted, Target, Vs Target
                    if isinstance(value, (int, float)):
                        cell.number_format = '£#,##0'
                        cell.alignment = currency_alignment
            
            current_row += 3  # Space between months
        
        # Set column widths
        column_widths = [18, 22, 24, 16, 16, 18, 18]
        if company.lower() == 'cnc':
            column_widths.append(8)  # C&C Apps
        column_widths.extend([18, 18, 20, 18, 14, 12])
        
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

    def _create_ytd_totals_sheet(self, workbook, data, company):
        """Create the YTD Totals sheet with Q3 and Q4 support"""
        ws = workbook.create_sheet("YTD Totals")
        
        # Styles (same as before)
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
        title_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=12)
        totals_font = Font(bold=True, size=10)
        center_alignment = Alignment(horizontal='center', vertical='center')
        currency_alignment = Alignment(horizontal='right', vertical='center')
        
        # Title and metadata (same as before)
        ws.merge_cells('A1:H1')  # Expanded for Q3/Q4 columns
        ws['A1'] = f"{data.get('team_name', '')} - YTD Totals Report"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_alignment
        
        ws['A2'] = f"Company: {company.upper()}"
        ws['A3'] = f"Period: {data.get('date_range', {}).get('start', '')} to {data.get('date_range', {}).get('end', '')}"
        ws['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        current_row = 6
        ytd_data = data.get('ytd_data', {})
        
        # Total Submitted section with Q3/Q4
        if 'total_submitted' in ytd_data:
            ws[f'A{current_row}'] = "TOTAL SUBMITTED"
            ws[f'A{current_row}'].font = section_font
            current_row += 1
            
            # Headers - NOW WITH Q3 AND Q4
            headers = ['Advisor', 'Quarter 1', 'Quarter 2', 'Quarter 3', 'Quarter 4', 'YTD Total', 'Yearly Target', 'Vs Target']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            current_row += 1
            
            # Data rows with Q3/Q4
            for item in ytd_data['total_submitted']:
                row_data = [
                    item.get('advisor', ''),
                    item.get('q1', 0),
                    item.get('q2', 0),
                    item.get('q3', 0),  # NEW
                    item.get('q4', 0),  # NEW
                    item.get('ytd_total', 0),
                    item.get('yearly_target', 0),
                    item.get('vs_target', 0)
                ]
                
                for col, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    
                    # Apply currency formatting to money columns (Q1-Q4, YTD, Target, Vs Target)
                    if col in [2, 3, 4, 5, 6, 7, 8]:
                        if isinstance(value, (int, float)):
                            cell.number_format = '£#,##0'
                            cell.alignment = currency_alignment
                
                current_row += 1
            
            current_row += 2
        
        # Appointments Completed section with Q3/Q4
        if 'appointments_completed' in ytd_data:
            ws[f'A{current_row}'] = "APPOINTMENTS COMPLETED"
            ws[f'A{current_row}'].font = section_font
            current_row += 1
            
            # Headers with Q3/Q4
            headers = ['Advisor', 'Quarter 1', 'Quarter 2', 'Quarter 3', 'Quarter 4', 'YTD Total']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            current_row += 1
            
            # Data rows with Q3/Q4
            for item in ytd_data['appointments_completed']:
                row_data = [
                    item.get('advisor', ''),
                    item.get('q1', 0),
                    item.get('q2', 0),
                    item.get('q3', 0),  # NEW
                    item.get('q4', 0),  # NEW
                    item.get('ytd_total', 0)
                ]
                
                for col, value in enumerate(row_data, start=1):
                    ws.cell(row=current_row, column=col, value=value)
                
                current_row += 1
            
            current_row += 2
        
        # Total Apps section with Q3/Q4
        if 'total_apps' in ytd_data:
            ws[f'A{current_row}'] = "TOTAL APPLICATIONS"
            ws[f'A{current_row}'].font = section_font
            current_row += 1
            
            # Headers with Q3/Q4
            headers = ['Advisor', 'Quarter 1', 'Quarter 2', 'Quarter 3', 'Quarter 4', 'YTD Total']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            current_row += 1
            
            # Data rows with Q3/Q4
            for item in ytd_data['total_apps']:
                row_data = [
                    item.get('advisor', ''),
                    item.get('q1', 0),
                    item.get('q2', 0),
                    item.get('q3', 0),  # NEW
                    item.get('q4', 0),  # NEW
                    item.get('ytd_total', 0)
                ]
                
                for col, value in enumerate(row_data, start=1):
                    ws.cell(row=current_row, column=col, value=value)
                
                current_row += 1
        
        # Set column widths for 8 columns
        column_widths = [20, 12, 12, 12, 12, 12, 14, 12]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width   
    def get_pipeline_summary(self, team_id):
        """Get pipeline summary with case stages breakdown"""
        try:
            user = self.get_current_user()
            if not user or not user.is_master:
                return jsonify({'error': 'Access denied'}), 403
            
            current_company = SessionManager.get_current_company(session)
            team = Team.query.filter_by(id=team_id, company=current_company).first()
            if not team:
                return jsonify({'error': 'Team not found'}), 404
            
            # Parse date parameters - FIXED to respect both start and end dates
            end_date_str = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
            start_date_str = request.args.get('start_date')
            
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            
            # Use provided start_date or default to start of year
            if start_date_str:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            else:
                start_date = datetime(end_date.year, 1, 1)
            
            team_members = team.members
            pipeline_data = []
            
            # Calculate total cases for percentage calculation
            total_team_cases = 0
            
            for member in team_members:
                # Get case stages breakdown for CURRENT COMPANY ONLY
                submissions = Submission.query.filter(
                    Submission.advisor_name == member.full_name,
                    Submission.company == current_company,
                    Submission.submission_date >= start_date,
                    Submission.submission_date <= end_date
                ).all()
                
                # Count by stages (customize these based on your actual stages)
                started = len([s for s in submissions if 'submitted' in s.business_type.lower()])
                fact_find = len([s for s in submissions if s.expected_proc > 0])  # Has value
                recommendation = len([s for s in submissions if s.expected_fee > 0])  # Has fee
                submitted_count = len(submissions)
                
                # Get exchange data from paid cases for CURRENT COMPANY ONLY
                exchanges = PaidCase.query.filter(
                    PaidCase.advisor_name == member.full_name,
                    PaidCase.company == current_company,
                    PaidCase.date_paid >= start_date,
                    PaidCase.date_paid <= end_date
                ).count()
                
                total_cases = started + fact_find + recommendation + submitted_count + exchanges
                total_team_cases += total_cases
                
                pipeline_data.append({
                    'advisor': member.full_name,
                    'started': started,
                    'fact_find': fact_find,
                    'recommendation': recommendation,
                    'submitted': submitted_count,
                    'exchange': exchanges,
                    'totals': total_cases,
                    'percent_of_total': 0  # Will calculate after we have total_team_cases
                })
            
            # Now calculate percentages correctly
            for item in pipeline_data:
                if total_team_cases > 0:
                    item['percent_of_total'] = round((item['totals'] / total_team_cases * 100), 1)
                else:
                    item['percent_of_total'] = 0.0
            
            # Calculate totals
            totals = {
                'started': sum(item['started'] for item in pipeline_data),
                'fact_find': sum(item['fact_find'] for item in pipeline_data),
                'recommendation': sum(item['recommendation'] for item in pipeline_data),
                'submitted': sum(item['submitted'] for item in pipeline_data),
                'exchange': sum(item['exchange'] for item in pipeline_data),
                'totals': sum(item['totals'] for item in pipeline_data)
            }
            
            # Calculate percentages for totals row
            total_all = totals['totals']
            percent_totals = {
                'started': round((totals['started'] / total_all * 100), 1) if total_all > 0 else 0,
                'fact_find': round((totals['fact_find'] / total_all * 100), 1) if total_all > 0 else 0,
                'recommendation': round((totals['recommendation'] / total_all * 100), 1) if total_all > 0 else 0,
                'submitted': round((totals['submitted'] / total_all * 100), 1) if total_all > 0 else 0,
                'exchange': round((totals['exchange'] / total_all * 100), 1) if total_all > 0 else 0
            }
            
            return jsonify({
                'success': True,
                'team_name': team.name,
                'company': current_company,
                'date_range': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                },
                'pipeline_data': pipeline_data,
                'totals': totals,
                'percent_totals': percent_totals
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500
    
    def test_calendly_debug(self):
        """Debug Calendly integration using existing methods"""
        try:
            print("🔍 Testing Calendly service...")
            
            # Test basic connection
            user_info = self.calendly_service.get_user_info()
            has_user_info = bool(user_info and 'resource' in user_info)
            
            # Get organization users
            org_users = self.calendly_service.get_organization_users()
            has_org_users = bool(org_users and 'collection' in org_users)
            
            # Test recent events
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)
            recent_events = self.calendly_service.get_scheduled_events(start_date, end_date, count=10)
            has_events = bool(recent_events and 'collection' in recent_events)
            
            # Get analytics data
            analytics_data = self.calendly_service.get_analytics_data_by_user(start_date, end_date)
            has_analytics = bool(analytics_data and 'users' in analytics_data)
            
            return jsonify({
                'calendly_status': '✅ Working' if has_user_info else '❌ Not Working',
                'has_token': bool(self.calendly_service.access_token),
                'token_length': len(self.calendly_service.access_token) if self.calendly_service.access_token else 0,
                'user_info': {
                    'available': has_user_info,
                    'name': user_info.get('resource', {}).get('name') if user_info else None,
                    'email': user_info.get('resource', {}).get('email') if user_info else None
                },
                'organization_users': {
                    'available': has_org_users,
                    'count': len(org_users.get('collection', [])) if org_users else 0
                },
                'recent_events': {
                    'available': has_events,
                    'count': len(recent_events.get('collection', [])) if recent_events else 0
                },
                'analytics_data': {
                    'available': has_analytics,
                    'users_count': len(analytics_data.get('users', {})) if analytics_data else 0
                }
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Debug failed: {str(e)}'})
    
    def test_calendly_emails(self):
        """Test what emails are available in Calendly"""
        try:
            print("📧 Getting available Calendly emails...")
            
            # Get analytics data for last 30 days
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
            
            analytics_data = self.calendly_service.get_analytics_data_by_user(start_date, end_date)
            
            if not analytics_data or 'users' not in analytics_data:
                return jsonify({
                    'success': False,
                    'error': 'No Calendly users data available',
                    'raw_response': analytics_data
                })
            
            # Extract user emails and info
            calendly_users = []
            for user_uri, user_data in analytics_data['users'].items():
                calendly_users.append({
                    'email': user_data.get('email'),
                    'name': user_data.get('name'),
                    'events_count': user_data.get('events_count', 0),
                    'uri': user_uri
                })
            
            return jsonify({
                'success': True,
                'calendly_users': calendly_users,
                'total_users': len(calendly_users),
                'date_range': f"{start_date.date()} to {end_date.date()}"
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Email test failed: {str(e)}'})
        
    def _get_real_calendly_data(self, member, start_date, end_date):
        """Get real Calendly data for a specific member and return (appointments_booked, appointments_completed)"""
        try:
            from app.services.calendly_service import CalendlyService
            calendly_service = CalendlyService()
            
            print(f"🗓️ Getting Calendly data for {member.full_name}")
            print(f"   Member email: {member.email}")
            print(f"   Date range: {start_date} to {end_date}")
            
            # Use get_events_for_user_email method
            events = calendly_service.get_events_for_user_email(member.email, start_date, end_date)
            
            if events:
                print(f"   Found {len(events)} total events")
                
                # Count events by status and time
                appointments_booked = 0
                appointments_completed = 0
                now = datetime.now()
                
                for event in events:
                    try:
                        start_time_str = event.get('start_time', '')
                        if start_time_str:
                            event_time = datetime.fromisoformat(start_time_str.replace('Z', '+00:00')).replace(tzinfo=None)
                            status = event.get('status', '').lower()
                            
                            # Count all events as "booked"
                            appointments_booked += 1
                            
                            # Count as "completed" if:
                            # 1. Event is in the past (before now)
                            # 2. Status is 'active' (means it happened) - NOT canceled
                            if event_time <= now and status == 'active':
                                appointments_completed += 1
                                
                            print(f"   Event: {event.get('name')} | {event_time.strftime('%Y-%m-%d')} | Status: {status} | Past: {event_time <= now}")
                            
                    except (ValueError, TypeError) as e:
                        print(f"   Error parsing event: {e}")
                        continue
                
                print(f"   ✅ {member.full_name}: {appointments_booked} booked, {appointments_completed} completed")
                return appointments_booked, appointments_completed
            else:
                print(f"   ⚠️ No events found for {member.full_name}")
                return 0, 0
                
        except Exception as e:
            print(f"   ❌ Error getting Calendly data for {member.full_name}: {e}")
            import traceback
            traceback.print_exc()
            return 0, 0
        

    def test_excel_download(self):
        """Test route to check Excel download functionality"""
        try:
            user = self.get_current_user()
            if not user or not user.is_master:
                return jsonify({'error': 'Access denied'}), 403
            
            # Get test parameters
            team_id = request.args.get('team_id', 1, type=int)
            
            print(f"Testing Excel download for team {team_id}")
            
            # Test the download functionality
            return self.download_ytd_excel(team_id)
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Test failed: {str(e)}'}), 500
