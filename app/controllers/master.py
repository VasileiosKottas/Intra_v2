"""
Master dashboard controller
"""

from flask import render_template, session, redirect, url_for, send_file, request, jsonify
from app.controllers.base import BaseController
from app.models.team import Team
from app.models.advisor import Advisor
from app.models.sync_log import SyncLog
from app.config.session import SessionManager
from app.config import config_manager
from app.models import db
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import io
from datetime import datetime

class MasterController(BaseController):
    """Handles master dashboard routes"""

    def register_routes(self):
        """Register master dashboard routes"""
        self.app.add_url_rule('/master', 'master.index', self.master_required(self.index))
        self.app.add_url_rule('/master/advisor/<int:advisor_id>', 'master.view_advisor', 
                            self.master_required(self.view_advisor_dashboard))
        self.app.add_url_rule('/master/team-performance-report', 'master.team_performance_report',
                            self.master_required(self.team_performance_report))
        # Add YTD Dashboard
        self.app.add_url_rule('/master/ytd-dashboard', 'master.ytd_dashboard',
                            self.master_required(self.ytd_dashboard))

    def index(self):
        """Master dashboard view"""
        user = self.get_current_user()
        if not user:
            session.clear()
            return redirect(url_for('auth.login'))
        
        current_company = SessionManager.get_current_company(session)
        teams = Team.query.filter_by(company=current_company).all()
        advisors = Advisor.query.filter_by(is_master=False).all()
        all_advisor_names = config_manager.get_advisor_names(current_company)
        recent_syncs = SyncLog.query.filter_by(company=current_company).order_by(SyncLog.sync_time.desc()).limit(10).all()
        
        company_config = SessionManager.get_company_config(session)
        return render_template('master.html', 
                             user=user, 
                             teams=teams, 
                             advisors=advisors, 
                             all_advisor_names=all_advisor_names,
                             recent_syncs=recent_syncs,
                             company_config=company_config)
    
    def ytd_dashboard(self):
        """Serve YTD dashboard page"""
        user = self.get_current_user()
        return render_template('master_ytd_dashboard.html', user=user)

    def view_advisor_dashboard(self, advisor_id):
        """Master view of advisor dashboard"""
        advisor = db.session.get(Advisor, advisor_id)
        if not advisor:
            return "Advisor not found", 404
        
        current_company = SessionManager.get_current_company(session)
        company_config = SessionManager.get_company_config(session)
        return render_template('advisor_view.html', 
                             advisor=advisor, 
                             company_config=company_config,
                             is_master_view=True,
                             current_company=current_company)
    
    def team_performance_report(self):
        """Serve team performance report page"""
        user = self.get_current_user()
        if not user or not user.is_master:
            return redirect(url_for('master.index'))
        
        return render_template('team_performance_report.html', user=user)
    
    def team_performance_excel(self, team_id):
        """Generate and download Excel report for team performance"""
        try:
            user = self.get_current_user()
            if not user or not user.is_master:
                return redirect(url_for('master.index'))
            
            from app.config.session import SessionManager
            from app.models.team import Team
            
            current_company = SessionManager.get_current_company(session)
            team = Team.query.filter_by(id=team_id, company=current_company).first()
            
            if not team:
                return jsonify({'error': 'Team not found'}), 404
            
            # Get month parameter
            month = request.args.get('month', datetime.now().strftime('%Y-%m'))
            
            # Generate Excel file
            excel_buffer = self._create_team_performance_excel(team, month, current_company)
            
            # Create filename
            month_name = datetime.strptime(month, '%Y-%m').strftime('%B_%Y')
            filename = f"{team.name}_{month_name}_Performance_Report.xlsx"
            
            return send_file(
                excel_buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    def _create_team_performance_excel(self, team, month, company):
        """Create Excel workbook with team performance data"""
        try:
            from app.config import config_manager
            from datetime import datetime, timedelta
            
            # Parse month
            report_date = datetime.strptime(month, '%Y-%m')
            start_date = report_date.replace(day=1)
            if report_date.month == 12:
                end_date = report_date.replace(year=report_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                end_date = report_date.replace(month=report_date.month + 1, day=1) - timedelta(days=1)
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{team.name} Performance"
            
            # Styles
            header_font = Font(bold=True, size=12, color='FFFFFF')
            header_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
            title_font = Font(bold=True, size=14)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells('A1:P1')
            ws['A1'] = f"{team.name} Performance Report - {report_date.strftime('%B %Y')}"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Add report metadata
            ws['A3'] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ws['A4'] = f"Company: {company.upper()}"
            ws['A5'] = f"Team Members: {len(team.members)}"
            
            # Headers
            headers = [
                'Advisor', 'Appointments Booked', 'Appointments Completed', 
                'Outbound Calls', 'Total Activity', 'Number of M Apps',
                'Insurance Apps', 'C&C Apps', 'Insurance Referrals', 
                'Submitted (Plus Fees)', 'Conversion %',
                'Target Activity', '% of Target Activity', 
                'Target Submitted', '% of Target Submitted'
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
            
            # Get team data
            team_members = team.members
            member_data = []
            
            for member in team_members:
                try:
                    # Get submission metrics
                    submission_metrics = member.calculate_metrics_for_period(
                        company, start_date, end_date,
                        config_manager.get_valid_business_types(company),
                        config_manager.get_valid_paid_case_types(company)
                    )
                    
                    # Placeholder data (same as in API)
                    appointments_booked = 15
                    appointments_completed = 12
                    outbound_calls = 85
                    total_activity = outbound_calls + appointments_completed
                    
                    # Real data from submissions
                    applications = submission_metrics.get('submissions_count', 0)
                    insurance_apps = self._count_business_type(member, start_date, end_date, company, 'insurance')
                    cnc_apps = self._count_business_type(member, start_date, end_date, company, 'conveyancing')
                    insurance_referrals = self._count_referrals(member, start_date, end_date, company, 'insurance')
                    
                    submitted_amount = submission_metrics.get('expected_proc', 0)
                    fees_amount = submission_metrics.get('expected_fee', 0)
                    submitted_plus_fees = submitted_amount + fees_amount
                    
                    total_apps = applications + insurance_apps + cnc_apps
                    conversion_rate = (total_apps / appointments_completed * 100) if appointments_completed > 0 else 0
                    
                    target_activity = 250
                    target_submitted = 50000
                    
                    member_data.append([
                        member.full_name,
                        appointments_booked,
                        appointments_completed,
                        outbound_calls,
                        total_activity,
                        applications,
                        insurance_apps,
                        cnc_apps,
                        insurance_referrals,
                        submitted_plus_fees,
                        f"{conversion_rate:.1f}%",
                        target_activity,
                        f"{(total_activity/target_activity*100):.0f}%" if target_activity > 0 else "N/A",
                        target_submitted,
                        f"{(submitted_plus_fees/target_submitted*100):.0f}%" if target_submitted > 0 else "N/A"
                    ])
                except Exception as e:
                    print(f"Error processing member {member.full_name}: {e}")
                    # Add placeholder row if member processing fails
                    member_data.append([
                        member.full_name, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "0%", 0, "N/A", 0, "N/A"
                    ])
            
            # Write member data
            for row, member in enumerate(member_data, 8):
                for col, value in enumerate(member, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.border = border
                    
                    # Format currency columns
                    if col == 11 or col == 15:  # Submitted amounts
                        if isinstance(value, (int, float)):
                            cell.number_format = '"£"#,##0.00'
                    
                    # Center align certain columns
                    if col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 13, 14, 16]:
                        cell.alignment = Alignment(horizontal='center')
            
            # Add totals row
            totals_row = len(member_data) + 8
            ws.cell(row=totals_row, column=1, value="TOTALS:").font = Font(bold=True)
            
            # Calculate totals
            try:
                for col in range(2, 17):
                    if col == 12:  # Conversion % - calculate team average
                        total_apps = sum(float(data[5]) + float(data[6]) + float(data[7]) for data in member_data)
                        total_completed = sum(float(data[2]) for data in member_data)
                        team_conversion = (total_apps / total_completed * 100) if total_completed > 0 else 0
                        ws.cell(row=totals_row, column=col, value=f"{team_conversion:.1f}%").font = Font(bold=True)
                    elif col in [14, 16]:  # Percentage columns - skip
                        continue
                    elif col == 11 or col == 15:  # Currency columns
                        total_val = sum(float(data[col-1]) if isinstance(data[col-1], (int, float)) else 0 for data in member_data)
                        cell = ws.cell(row=totals_row, column=col, value=total_val)
                        cell.font = Font(bold=True)
                        cell.number_format = '"£"#,##0.00'
                    else:
                        total_val = sum(float(data[col-1]) if isinstance(data[col-1], (int, float)) else 0 for data in member_data)
                        ws.cell(row=totals_row, column=col, value=total_val).font = Font(bold=True)
            except Exception as e:
                print(f"Error calculating totals: {e}")
            
            # Auto-adjust column widths - Fixed version
            for col_num in range(1, 17):  # Columns A to P
                column_letter = openpyxl.utils.get_column_letter(col_num)
                max_length = 0
                
                # Check all cells in this column
                for row_num in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                # Set column width
                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add data source notes
            notes_row = totals_row + 3
            ws[f'A{notes_row}'] = "Data Sources:"
            ws[f'A{notes_row}'].font = Font(bold=True)
            ws[f'A{notes_row+1}'] = "• Submission data: Live from database"
            ws[f'A{notes_row+2}'] = "• Appointments: Placeholder data (pending Calendly integration)"
            ws[f'A{notes_row+3}'] = "• Outbound calls: Placeholder data (pending ALTOS integration)"
            
            # Save to buffer
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except Exception as e:
            print(f"Error creating Excel file: {e}")
            # Create simple error workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = f"Error generating report for {team.name}: {str(e)}"
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            return excel_buffer
        
    def _count_business_type(self, advisor, start_date, end_date, company, business_type):
        """Count submissions by business type"""
        from app.models.submission import Submission
        
        return Submission.query.filter(
            Submission.advisor_name == advisor.full_name,
            Submission.company == company,
            Submission.submission_date >= start_date,
            Submission.submission_date <= end_date,
            Submission.business_type.ilike(f'%{business_type}%')
        ).count()
    
    def _count_referrals(self, advisor, start_date, end_date, company, referral_type):
        """Count referrals by type"""
        from app.models.submission import Submission
        
        if referral_type == 'insurance':
            return Submission.query.filter(
                Submission.advisor_name == advisor.full_name,
                Submission.company == company,
                Submission.submission_date >= start_date,
                Submission.submission_date <= end_date,
                Submission.business_type == 'Referral',
                Submission.referral_to.ilike('%insurance%')
            ).count()
        else:
            return Submission.query.filter(
                Submission.advisor_name == advisor.full_name,
                Submission.company == company,
                Submission.submission_date >= start_date,
                Submission.submission_date <= end_date,
                Submission.business_type == 'Referral',
                ~Submission.referral_to.ilike('%insurance%')
            ).count()